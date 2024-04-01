"""
    Author: AaronTook (https://AaronTook.github.io)
    File Last Modified: 3/31/2024
    Project Name: Staff Scheduler
    File Name: wrapper.py
"""

# Python Standard Library module imports.
import traceback

# Third-party module imports.
import openpyxl

# Project module imports.
from utils import *

class SchedulerApplication():
    """ Create objects to wrap the utility functions used in creating a schedule. """
    def __init__(self, input_filepath, output_filepath, debug_mode=False):
        """ Initialize the SchedulerApplication object's default attributes. """
        self.in_fp = input_filepath
        self.out_fp = output_filepath
        self.debug = debug_mode
        self.guards_per_shift = 5
        self.managers_per_shift = 1
        self.attendants_per_shift = 1
        self.max_consecutive_shifts = 3
       
    def set_per_shift(self, guards=None, managers=None, attendants=None):
        """ Setter for any of the _per_shift attributes. """
        if guards != None:
            if isinstance(guards, int):
                self.guards_per_shift = guards
            else:
                print("Error: SchedulerApplication.set_per_shift() argument \"guards\" only accepts values of type <class=\"int\">")
        if managers != None:
            if isinstance(managers, int):
                self.managers_per_shift = managers
            else:
                print("Error: SchedulerApplication.set_per_shift() argument \"managers\" only accepts values of type <class=\"int\">")
        if attendants != None:
            if isinstance(attendants, int):
                self.attendants_per_shift = attendants
            else:
                print("Error: SchedulerApplication.set_per_shift() argument \"attendants\" only accepts values of type <class=\"int\">")
    def set_input_file(self, new_in_filepath):
        """ Setter for in_fp attribute. """
        self.in_fp = new_in_filepath
    def set_output_file(self, new_out_filepath):
        """ Setter for out_fp attribute. """
        self.out_fp = new_out_filepath
    def set_max_consecutive_shifts(self, max_consecutive_shifts):
        """ Setter for max_consecutive_shifts attribute. """
        self.max_consecutive_shifts = max_consecutive_shifts
    def get_per_shift(self):
        """ Getter for _per_shift attributes. """
        print(f"Staff per shift by staff position:\nGuards: {self.guards_per_shift}\nManagers: {self.managers_per_shift}\nPool Attendants: {self.attendants_per_shift}")
        return self.guards_per_shift, self.managers_per_shift, self.attendants_per_shift
    def get_input_file(self):
        """ Getter for in_fp attribute. """
        print(f"Input file location: {self.in_fp}")
        return self.in_fp
    def get_output_file(self):
        """ Getter for out_fp attribute. """
        print(f"Output file location: {self.out_fp}")
        return self.out_fp
    def get_max_consecutive_shifts(self):
        """ Getter for max_consecutive_shifts attribute. """
        return self.max_consecutive_shifts

    def run(self):
        opened_in_fp = False
        try:
            wb = openpyxl.load_workbook(self.in_fp)
            opened_in_fp = True
            lg = load_wb_ws(wb)
            pa = load_wb_ws(wb, "Pool Attendants")
            mg = load_wb_ws(wb, "Managers")
            
            # Create the workbook and set up the worksheet.
            export_wb = openpyxl.Workbook()
            export_ws = export_wb.active
        
            guard_schedule = parse_data_struct(lg, self.guards_per_shift, max_consecutive_shifts = self.max_consecutive_shifts, debug_mode = self.debug)
            attendant_schedule = parse_data_struct(pa, self.attendants_per_shift, max_consecutive_shifts = self.max_consecutive_shifts, debug_mode = self.debug)
            manager_schedule = parse_data_struct(mg, self.managers_per_shift, max_consecutive_shifts = self.max_consecutive_shifts, debug_mode = self.debug)
            
            ws_1 = export_wb.active
            guard_ws = export_data_to_xlsx(guard_schedule, ws_1, "Lifeguards", debug_mode = self.debug)
            
            ws_2 = export_wb.create_sheet()
            attendant_ws = export_data_to_xlsx(attendant_schedule, ws_2, "Pool Attendants", debug_mode = self.debug)
            
            ws_3 = export_wb.create_sheet()
            manager_ws = export_data_to_xlsx(manager_schedule, ws_3, "Managers", debug_mode = self.debug)
            
            export_wb.save(self.out_fp) # Save the document.
            return self.out_fp
        except PermissionError:
            if opened_in_fp:
                print(f"Error: could not open the output file \"{self.out_fp}\" for writing. Please close any other application using the file before running.")
            else:
                print(f"Error: could not open the input file \"{self.in_fp}\" for parsing. Please close any other application using the file before running.")
            return False
        except Exception as e:
            print(f"Error: application encountered unexpected crash. Full error message: \n{traceback.format_exc()}")
            return False

