"""
    Author: AaronTook (https://AaronTook.github.io)
    File Last Modified: 3/31/2024
    Project Name: Staff Scheduler
    File Name: utils.py
"""

# Python Standard Library module imports.
from random import choice
from tkinter import *
from tkinter import filedialog
import os

# Third-party module imports.
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import PatternFill, Alignment


def gui_get_file(initial_directory="", limit_filetypes=[]):
    """ Open file explorer (using tkinter) to select a file. """
    root = Tk()
    root.withdraw()
    complete_file_path = filedialog.askopenfilename(title="PyPersonalVault - File Select", initialdir = os.getcwd() + "/" + initial_directory, filetypes = limit_filetypes)
    root.destroy()
    # Extract the path and filename and return those strings.
    file_path, file_name = os.path.split(complete_file_path)
    return complete_file_path, file_name

def pick_staff_from_randomizer(staff_list, staff_needed, backup_staff=[], debug_mode = False):
    """ Efficiently select staff randomly using one of two methods. """
    scheduled_staff = []
    num_available = len(staff_list)
    num_backup = len(backup_staff)
    if num_available >= staff_needed: # Backup staff is not needed.
        if num_available < staff_needed * 2: # Randomly pick who will not be scheduled.
            for i in range(num_available - staff_needed):
                staff_member = choice(staff_list)
                staff_list.remove(staff_member)
            scheduled_staff = [staff_member + "<mediumCertainty>" for staff_member in staff_list]
        else: # Randomly pick who will be scheduled.
            for i in range(staff_needed):
                staff_member = choice(staff_list)
                scheduled_staff.append(staff_member + "<mediumCertainty>")
                staff_list.remove(staff_member)
    else: # Backup staff is needed.
        num_backup_needed = staff_needed-num_available
        if debug_mode:
            print("Backup staff needed.")
            print(staff_list, backup_staff)
            print(f"Backups needed: {num_backup_needed}")
        sorted_backup_staff = sorted(backup_staff, key=lambda d: d["consecutive_shifts"])
        least_consecutive = sorted_backup_staff[0]["consecutive_shifts"]
        most_consecutive = sorted_backup_staff[-1]["consecutive_shifts"]
        
        # Split sorted_backup_staff into multiple lists by consecutive_shifts count.
        split_keys = [f"{consecutive_i + least_consecutive}_consecutive_shifts" for consecutive_i in range(most_consecutive)]
        split_sorted_backup_staff = { f"{consecutive_i + least_consecutive}_consecutive_shifts": [] for consecutive_i in range(most_consecutive + 1 - least_consecutive)} # Create the empty data structure to hold lists of staff who have worked the same number of consecutive shifts.
        for sorted_backup in sorted_backup_staff: # Populate split_sorted_backup_staff.
            split_sorted_backup_staff[f"{sorted_backup['consecutive_shifts']}_consecutive_shifts"].append(sorted_backup["staff_member"])
        iterable_backups = [x_consecutive_shifts for x_consecutive_shifts in split_sorted_backup_staff.values()] # Get only the values of the dictionary "split_sorted_backup_staff".
        if debug_mode:
            print(sorted_backup_staff)
            print(split_sorted_backup_staff)
        
        # Schedule all non-backup staff first.
        for staff_member in staff_list:
            scheduled_staff.append(staff_member + "<mediumCertainty>")
        
        num_still_needed = num_backup_needed
        while num_still_needed > 0: #Continue selection from backup staff until all the staffing needs are met.
            if num_still_needed >= len(iterable_backups[0]): # No randomization, pick all staff who worked the given number of consecutive shifts.
                for staff_with_x_shifts in iterable_backups[0]:
                    if debug_mode:
                        print(f"Adding \"{staff_with_x_shifts}\" to the shift.")
                    scheduled_staff.append(staff_with_x_shifts + "<backupCertainty>")
                num_still_needed -= len(iterable_backups[0]) # Decrease the number of staff still needed for the shift.
                iterable_backups.pop(0) # Remove all the staff members who worked the given number of consecutive shifts.
            else: # Pick only one staff member who worked the given number of consecutive shifts.
                staff_member = choice(iterable_backups[0])
                if debug_mode:
                    print(f"Adding \"{staff_member}\" to the shift by random selection.")
                iterable_backups[0].remove(staff_member)
                scheduled_staff.append(staff_member + "<backupCertainty>") # Add the selected staff member.
                num_still_needed -= 1 # Decrease the number of staff still needed for the shift.

    return scheduled_staff

def load_wb_ws(wb_obj, ws_name="Lifeguards"):
    """ Load the data from the workbook wb_obj representing staff availability. """
    data_struct = {}
    # Load the correct worksheet.
    ws = wb_obj[ws_name]
    header = None
    for col in ws.iter_cols(): # Iterate through all the workbook data.
        if col[0].value != None:
            header = col[0].value
        data_struct[header + "\n" + col[1].value] = [(cell.value).lower().strip().title() for cell in col[2:] if cell.value!=None]
    return data_struct

def parse_data_struct(data_struct, staff_needed, max_consecutive_shifts=3, debug_mode = False):
    """ Create a schedule from the data_struct containing staff availability. """
    # Separate the data from data_struct.
    data_keys = [key for key in data_struct.keys()]
    data_values = [value for value in data_struct.values()]
    
    # Create the new schedule.
    new_schedule = {}
    # Iterate through each shift.
    for item_index in range(len(data_struct.keys())):
        item_key = data_keys[item_index]
        item_value = data_values[item_index]
        available_staff = len(item_value)
        if debug_mode:
            print(f"Processing shift {item_key.replace('\n','')} with available staff {item_value}")
        # Analyze the shift based on the number of staff available.
        if available_staff < staff_needed: # Less staff than is needed (open shifts available).
            staff_list = [staff_member + "<highCertainty>" for staff_member in item_value]
            for i in range(staff_needed - available_staff): #  Add blank staff members for each available shift.
                staff_list.append("<lowCertainty>")
            new_schedule[item_key] = staff_list
            
        elif available_staff == staff_needed: # The exact number of staff needed are available (no assignment logic needed, no open shifts).
            staff_list = [staff_member + "<highCertainty>" for staff_member in item_value]
            new_schedule[item_key] = staff_list
            
        else: # More staff is available than is needed (assignment logic needed, no open shifts).
            available_staff_members = item_value
            current_shift_i = item_index
            backup_staff = []
            if current_shift_i >= max_consecutive_shifts: # The current shift is at least the first shift after the maximum number of consecutive shifts.
                schedule_values = [shift for shift in new_schedule.values()]
                previous_shifts = [schedule_values[current_shift_i - (1 + i)] for i in range(current_shift_i)] # Get a list of all previous shifts.
                for available_staff_member in item_value: # Logic for each staff member.
                    consecutive_shifts_worked = 0
                    for previous_shift in previous_shifts: # Calculate the available staff member's number of consecutive shifts.
                        if available_staff_member in [previous_shift_staff_member.replace("<highCertainty>","").replace("<mediumCertainty>","").replace("<lowCertainty>","").replace("<backupCertainty>","") for previous_shift_staff_member in previous_shift]:
                            consecutive_shifts_worked += 1
                        else: # The staff member did not work the shift, so stop counting consecutive shifts.
                            break
                    if consecutive_shifts_worked >= max_consecutive_shifts: # Their maximum ideal number of shifts has been reached.
                        if debug_mode:
                            print(f"\"{available_staff_member}\" has maxed out and is now backup for {item_key.replace('\n', '')}")
                        backup_staff.append({"staff_member": available_staff_member, "consecutive_shifts": consecutive_shifts_worked}) # Designate them as backup staff (represent both staff member and number of consecutive shifts worked.
                for backup_staff_member in [backup_staff_item["staff_member"] for backup_staff_item in backup_staff]: # This must be done after all the backup staff has been selected, otherwise the logic skips the user immediately after any user designated as backup.
                    available_staff_members.remove(backup_staff_member)
                if debug_mode:
                    print(available_staff_members)
                staff_list = pick_staff_from_randomizer(available_staff_members, staff_needed, backup_staff, debug_mode=debug_mode) # Pick the staff to schedule.
                if debug_mode:
                    print(staff_list)
            
            else: # There haven't been enough consecutive shifts yet for staff to have reached the maximum ideal consecutive shift number.
                staff_list = pick_staff_from_randomizer(available_staff_members, staff_needed, debug_mode = debug_mode)
            
            new_schedule[item_key] = [staff_member for staff_member in staff_list] # Add the staff for the shift to the new schedule.
                
    return new_schedule

def export_data_to_xlsx(data_struct, ws_object, ws_name="Lifeguards", debug_mode=False):
    """ Create .xlsx file from the data_struct. """
    # Rename the worksheet
    export_ws = ws_object
    export_ws.title = ws_name
    
    data_keys = [key for key in data_struct.keys()]
    data_values = [value for value in data_struct.values()]
    
    # Define constant workbook values.
    start_col_i = 1 # Start shifts in column B.
    start_row_i = 2 # Start shifts in row 3.
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid') # Defines fill colors.
    orangeFill = PatternFill(start_color='00FF9900', end_color='00FF9900', fill_type='solid') 
    darkGreenFill = PatternFill(start_color='00008000', end_color='00008000', fill_type='solid')
    lightGreenFill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
    
    # Iterate through shifts.
    for i in range(len(data_keys)): # Add each shift header.
        current_col = get_column_letter(start_col_i + i + 1)
        key_parts = data_keys[i].split("\n")
        export_ws[f"{current_col}{start_row_i}"] = key_parts[1]
        export_ws[f"{current_col}{start_row_i - 1}"] = key_parts[0]
        if (export_ws[f"{get_column_letter(start_col_i + i)}{start_row_i-1}"]).value == key_parts[0]:
            export_ws.merge_cells(f"{get_column_letter(start_col_i + i)}{start_row_i-1}:{current_col}{start_row_i-1}")
        
        # Iterate through staff list for each shift.
        for j in range(len(data_values[i])):
            # Add each staff member.
            current_row = start_row_i + j + 1
            current_cell_data = data_values[i][j]
            
            # Set fill color based on scheduling certainty for each shift.
            certainty = ""
            cell_color = None
            if current_cell_data.endswith("<highCertainty>"): # High certainty tag. Mark dark green.
                certainty = "high"
                current_cell_data = current_cell_data[:-15]
                cell_color = darkGreenFill
            elif current_cell_data.endswith("<mediumCertainty>"): # Medium certainty tag. Mark light green.
                certainty = "medium"
                current_cell_data = current_cell_data[:-17]
                cell_color = lightGreenFill
            elif current_cell_data.endswith("<lowCertainty>"): # Low certainty tag. Mark red.
                certainty = "low"
                current_cell_data = current_cell_data[:-14]
                cell_color = redFill
            elif current_cell_data.endswith("<backupCertainty>"): # Backup certainty tag. Mark orange.
                certainty = "backup"
                current_cell_data = current_cell_data[:-17]
                cell_color = orangeFill
            else: # Invalid certainty tag. Mark red.
                certainty = "red"
                cell_color = redFill
            
            export_ws[f"{current_col}{current_row}"] = current_cell_data
            export_ws[f"{current_col}{current_row}"].fill = cell_color
            
     # Adjust column widths.
    dim_holder = DimensionHolder(worksheet=export_ws)
    for col in range(export_ws.min_column, export_ws.max_column + 1): # Calculate the width of each column so all data fits.
        dim_holder[get_column_letter(col)] = ColumnDimension(export_ws, min=col, max=col, width=20)
    export_ws.column_dimensions = dim_holder # Set the dimensions.
    
    # Set the worksheet's first column (row labels).
    staff_needed = len(data_values[-1]) # Calculate the number of staff per shift.
    start_col = get_column_letter(start_col_i)
    export_ws[f"{start_col}{start_row_i - 1}"] = "Day" # R1 = Day
    export_ws[f"{start_col}{start_row_i}"] = "Shift" # R2 = Shift
    for i in range(staff_needed): # R3...R3+i = Staff
        export_ws[f"{start_col}{start_row_i + i + 1}"] = "Staff"
    export_ws.merge_cells(f"{start_col}{start_row_i + 1}:{start_col}{start_row_i + i + 1}") # Merge the labels for all "Staff" rows.
    export_ws[f"{start_col}{start_row_i - 1}"].alignment = Alignment(horizontal = "center", vertical="center") # Center align all of the first column cells.
    export_ws[f"{start_col}{start_row_i}"].alignment = Alignment(horizontal = "center", vertical="center")
    export_ws[f"{start_col}{start_row_i + 1}"].alignment = Alignment(horizontal = "center", vertical="center")
    export_ws.freeze_panes = "B2" # Freeze the first two rows and first column.
    
    return export_ws